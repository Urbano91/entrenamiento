"""Vista de Club: altas iniciales y lectura, nunca edición del trabajo técnico."""

from __future__ import annotations

from datetime import date, time, timedelta
from io import BytesIO
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from starlette.responses import StreamingResponse

from app.api.auth import get_current_user
from app.db.database import get_db
from app.models.models import (
    Club,
    CoachAssignment,
    Entrenamiento,
    ExerciseOwnership,
    Partido,
    PerfilEntrenador,
    PlanificacionDiaria,
    Temporada,
    UserAccount,
    Usuario,
)
from app.services.permissions import require_club
from app.services.trainer_accounts import (
    create_trainer_account, generate_trainer_credentials,
)
from app.services.trainer_colors import trainer_color


router = APIRouter(prefix="/api/club", tags=["Club"])


class ClubTrainerOut(BaseModel):
    assignment_id: int
    user_id: int
    usuario: str
    nombre: str
    apellidos: str
    club: str
    temporada_id: int
    temporada: str
    categoria: str
    puesto: str
    parent_coach_assignment_id: Optional[int] = None
    onboarding_complete: bool
    exercise_count: int
    training_count: int
    match_count: int
    color: str
    visible: bool
    provisional_password: Optional[str] = None


class ClubTrainerVisibilityIn(BaseModel):
    visible: bool


class ClubTrainerCreate(BaseModel):
    nombre: str = Field(min_length=1, max_length=120)
    apellidos: str = Field(min_length=1, max_length=180)
    temporada_id: int
    categoria: str = Field(min_length=1, max_length=120)
    puesto: str = Field(min_length=1, max_length=120)
    parent_coach_assignment_id: Optional[int] = None


def assignment_payload(db: Session, assignment: CoachAssignment) -> dict:
    profile = db.query(PerfilEntrenador).filter(
        PerfilEntrenador.usuario_id == assignment.coach_user_id
    ).one()
    account = db.get(UserAccount, assignment.coach_user_id)
    exercise_count = db.query(ExerciseOwnership).filter(
        ExerciseOwnership.created_by_user_id == assignment.coach_user_id,
        ExerciseOwnership.deleted_at.is_(None),
    ).count()
    training_count = db.query(Entrenamiento).filter(
        Entrenamiento.usuario_id == assignment.coach_user_id,
        Entrenamiento.temporada_id == assignment.temporada_id,
    ).count()
    match_count = db.query(Partido).filter(
        Partido.usuario_id == assignment.coach_user_id,
        Partido.temporada_id == assignment.temporada_id,
    ).count()
    return {
        "assignment_id": assignment.id,
        "user_id": assignment.coach_user_id,
        "usuario": assignment.coach.usuario,
        "nombre": profile.nombre,
        "apellidos": profile.apellidos,
        "club": assignment.club.nombre,
        "temporada_id": assignment.temporada_id,
        "temporada": assignment.temporada.nombre,
        "categoria": assignment.category.nombre,
        "puesto": assignment.puesto,
        "parent_coach_assignment_id": assignment.parent_coach_assignment_id,
        "onboarding_complete": bool(account and account.onboarding_complete),
        "exercise_count": exercise_count,
        "training_count": training_count,
        "match_count": match_count,
        "color": trainer_color(assignment.coach_user_id),
        "visible": assignment.visible_in_club,
        "provisional_password": None,
    }


@router.get("", response_model=dict)
def get_club(current_user: Usuario = Depends(get_current_user), db: Session = Depends(get_db)):
    club = require_club(db, current_user)
    return {"id": club.id, "nombre": club.nombre, "owner_user_id": club.owner_user_id}


@router.get("/entrenadores", response_model=list[ClubTrainerOut])
def list_trainers(
    temporada_id: Optional[int] = None,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    club = require_club(db, current_user)
    query = db.query(CoachAssignment).filter(
        CoachAssignment.club_id == club.id, CoachAssignment.active.is_(True)
    )
    if temporada_id is not None:
        query = query.filter(CoachAssignment.temporada_id == temporada_id)
    return [assignment_payload(db, item) for item in query.order_by(CoachAssignment.id).all()]


@router.post("/entrenadores", response_model=ClubTrainerOut, status_code=201)
def create_club_trainer(
    data: ClubTrainerCreate,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    club = require_club(db, current_user)
    season = db.get(Temporada, data.temporada_id)
    if season is None:
        raise HTTPException(status_code=422, detail="Temporada no válida")

    if data.puesto == "Entrenador":
        if data.parent_coach_assignment_id is not None:
            raise HTTPException(
                status_code=422,
                detail="Un entrenador no debe depender de otro entrenador",
            )
    else:
        if data.parent_coach_assignment_id is None:
            raise HTTPException(
                status_code=422,
                detail="Debes indicar de qué entrenador depende este miembro del cuerpo técnico",
            )

        parent_assignment = db.query(CoachAssignment).filter(
            CoachAssignment.id == data.parent_coach_assignment_id,
            CoachAssignment.club_id == club.id,
            CoachAssignment.temporada_id == data.temporada_id,
            CoachAssignment.puesto == "Entrenador",
            CoachAssignment.active.is_(True),
        ).one_or_none()

        if parent_assignment is None:
            raise HTTPException(
                status_code=422,
                detail="El entrenador seleccionado no es válido para este club y temporada",
            )

    credentials = generate_trainer_credentials(db, data.nombre, data.apellidos)
    user = create_trainer_account(
        db,
        nombre=data.nombre,
        apellidos=data.apellidos,
        usuario=credentials.usuario,
        password_provisional=credentials.password_provisional,
        club=club,
        season=season,
        category_name=data.categoria,
        puesto=data.puesto,
        parent_coach_assignment_id=data.parent_coach_assignment_id,
    )
    assignment = db.query(CoachAssignment).filter(
        CoachAssignment.coach_user_id == user.id,
        CoachAssignment.club_id == club.id,
        CoachAssignment.temporada_id == season.id,
    ).one()
    payload = assignment_payload(db, assignment)
    payload["provisional_password"] = credentials.password_provisional
    return payload


def _coordination_data(
    db: Session,
    club: Club,
    temporada_id: int,
    selected_coach_user_id: Optional[int],
    desde: date,
    hasta: date,
) -> dict:
    if desde > hasta:
        raise HTTPException(status_code=422, detail="El rango de fechas no es válido")
    assignments = db.query(CoachAssignment).filter(
        CoachAssignment.club_id == club.id,
        CoachAssignment.temporada_id == temporada_id,
        CoachAssignment.active.is_(True),
        CoachAssignment.visible_in_club.is_(True),
        CoachAssignment.puesto == "Entrenador",
    ).order_by(CoachAssignment.id).all()
    if selected_coach_user_id is not None:
        assignments = [
            item for item in assignments
            if item.coach_user_id == selected_coach_user_id
        ]
        if not assignments:
            raise HTTPException(
                status_code=422,
                detail="El entrenador no está asignado al club en esta temporada",
            )
    trainer_ids = list(dict.fromkeys(item.coach_user_id for item in assignments))
    profiles = {
        item.usuario_id: item for item in db.query(PerfilEntrenador).filter(
            PerfilEntrenador.usuario_id.in_(trainer_ids or [-1])
        ).all()
    }
    trainers = []
    for item in assignments:
        profile = profiles.get(item.coach_user_id)
        trainers.append({
            "user_id": item.coach_user_id,
            "nombre": item.coach.usuario,
            "display_name": (
                f"{profile.nombre} {profile.apellidos}" if profile
                else item.coach.usuario
            ),
            "categoria": item.category.nombre,
            "color": trainer_color(item.coach_user_id),
        })

    activities = []
    trainings = db.query(Entrenamiento).filter(
        Entrenamiento.usuario_id.in_(trainer_ids or [-1]),
        Entrenamiento.temporada_id == temporada_id,
        Entrenamiento.fecha >= desde,
        Entrenamiento.fecha <= hasta,
    ).all()
    for training in trainings:
        assignment = next(
            item for item in assignments
            if item.coach_user_id == training.usuario_id
        )
        profile = profiles.get(training.usuario_id)
        activities.append({
            "type": "ENTRENAMIENTO",
            "id": training.id,
            "fecha": training.fecha.isoformat(),
            "hora": (
                training.hora.isoformat(timespec="minutes")
                if training.hora else None
            ),
            "trainer_user_id": training.usuario_id,
            "trainer": (
                f"{profile.nombre} {profile.apellidos}" if profile
                else assignment.coach.usuario
            ),
            "categoria": assignment.category.nombre,
            "title": training.nombre,
            "duration": training.duracion_minutos,
            "objective": training.objetivo_principal,
            "notes": training.observaciones,
            "color": trainer_color(training.usuario_id),
            "exercises": [
                {"id": rel.ejercicio.id, "nombre": rel.ejercicio.nombre}
                for rel in training.ejercicios_rel
            ],
        })
    matches = db.query(Partido).filter(
        Partido.usuario_id.in_(trainer_ids or [-1]),
        Partido.temporada_id == temporada_id,
        Partido.fecha >= desde,
        Partido.fecha <= hasta,
    ).all()
    for match in matches:
        assignment = next(
            item for item in assignments if item.coach_user_id == match.usuario_id
        )
        profile = profiles.get(match.usuario_id)
        activities.append({
            "type": "PARTIDO",
            "id": match.id,
            "fecha": match.fecha.isoformat(),
            "hora": match.hora.isoformat(timespec="minutes") if match.hora else None,
            "trainer_user_id": match.usuario_id,
            "trainer": (
                f"{profile.nombre} {profile.apellidos}" if profile
                else assignment.coach.usuario
            ),
            "categoria": assignment.category.nombre,
            "title": f"Partido vs {match.rival}",
            "duration": None,
            "objective": match.observaciones,
            "notes": match.observaciones,
            "color": trainer_color(match.usuario_id),
            "exercises": [],
        })
    plans = db.query(PlanificacionDiaria).filter(
        PlanificacionDiaria.usuario_id.in_(trainer_ids or [-1]),
        PlanificacionDiaria.temporada_id == temporada_id,
        PlanificacionDiaria.fecha >= desde,
        PlanificacionDiaria.fecha <= hasta,
    ).all()
    return {
        "club": {"id": club.id, "nombre": club.nombre},
        "temporada_id": temporada_id,
        "trainers": trainers,
        "activities": sorted(
            activities,
            key=lambda item: (
                item["fecha"], item["hora"] or "", item["trainer"].casefold()
            ),
        ),
        "planning": [{
            "trainer_user_id": item.usuario_id,
            "fecha": item.fecha.isoformat(),
            "trainer": (
                f"{profiles[item.usuario_id].nombre} {profiles[item.usuario_id].apellidos}"
                if item.usuario_id in profiles else str(item.usuario_id)
            ),
            "note": item.nota,
            "color": trainer_color(item.usuario_id),
            "documents": [document.nombre_original for document in item.documentos],
        } for item in plans],
    }



@router.get("/coordination")
def coordination(
    temporada_id: int,
    coach_user_id: Optional[int] = None,
    trainer_id: Optional[int] = None,
    desde: date = Query(default_factory=lambda: date.today() - timedelta(days=15)),
    hasta: date = Query(default_factory=lambda: date.today() + timedelta(days=45)),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    club = require_club(db, current_user)
    if (
        coach_user_id is not None
        and trainer_id is not None
        and coach_user_id != trainer_id
    ):
        raise HTTPException(status_code=422, detail="Filtro de entrenador incoherente")
    selected_coach_user_id = (
        coach_user_id if coach_user_id is not None else trainer_id
    )
    return _coordination_data(
        db, club, temporada_id, selected_coach_user_id,
        desde, hasta,
    )


def _light_fill(color: str) -> str:
    channels = [int(color[index:index + 2], 16) for index in (1, 3, 5)]
    return "".join(f"{round(value + (255 - value) * 0.78):02X}" for value in channels)


def _coordination_workbook(data: dict, season_name: str, desde: date, hasta: date) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calendario"
    sheet.merge_cells("A1:G1")
    sheet["A1"] = f"Calendario de coordinación · {data['club']['nombre']}"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="center")
    selected = data["trainers"]
    sheet["A2"] = "Temporada"
    sheet["B2"] = season_name
    sheet["C2"] = "Periodo"
    sheet["D2"] = f"{desde.strftime('%d/%m/%Y')} - {hasta.strftime('%d/%m/%Y')}"
    sheet["E2"] = "Entrenadores"
    sheet["F2"] = selected[0]["display_name"] if len(selected) == 1 else "Todos"

    headers = [
        "Fecha", "Hora", "Entrenador", "Categoría", "Tipo actividad",
        "Nombre", "Observaciones/notas",
    ]
    for column, value in enumerate(headers, 1):
        cell = sheet.cell(row=4, column=column, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24547A")
        cell.alignment = Alignment(horizontal="center")

    planning_notes = {
        (item["trainer_user_id"], item["fecha"]): item.get("note")
        for item in data["planning"] if item.get("note")
    }
    for row_number, activity in enumerate(data["activities"], 5):
        notes = [activity.get("notes"), planning_notes.get(
            (activity["trainer_user_id"], activity["fecha"])
        )]
        values = [
            date.fromisoformat(activity["fecha"]),
            time.fromisoformat(activity["hora"]) if activity.get("hora") else None,
            activity["trainer"],
            activity["categoria"],
            "Entrenamiento" if activity["type"] == "ENTRENAMIENTO" else "Partido",
            activity["title"],
            " · ".join(dict.fromkeys(item for item in notes if item)),
        ]
        fill = PatternFill("solid", fgColor=_light_fill(activity["color"]))
        for column, value in enumerate(values, 1):
            cell = sheet.cell(row=row_number, column=column, value=value)
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=column in (6, 7))
        sheet.cell(row=row_number, column=1).number_format = "dd/mm/yyyy"
        sheet.cell(row=row_number, column=2).number_format = "hh:mm"

    last_row = max(4, sheet.max_row)
    sheet.auto_filter.ref = f"A4:G{last_row}"
    sheet.freeze_panes = "A5"
    widths = [13, 10, 24, 18, 18, 34, 44]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[1].height = 27

    legend = workbook.create_sheet("Leyenda")
    legend.append(["Color", "Entrenador", "Categoría"])
    for cell in legend[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24547A")
    seen = set()
    for trainer in data["trainers"]:
        key = (trainer["user_id"], trainer["categoria"])
        if key in seen:
            continue
        seen.add(key)
        legend.append(["", trainer["display_name"], trainer["categoria"]])
        legend.cell(legend.max_row, 1).fill = PatternFill(
            "solid", fgColor=trainer["color"].removeprefix("#")
        )
    legend.column_dimensions["A"].width = 12
    legend.column_dimensions["B"].width = 28
    legend.column_dimensions["C"].width = 22
    legend.freeze_panes = "A2"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@router.get("/coordination/export.xlsx")
def export_coordination(
    temporada_id: int,
    coach_user_id: Optional[int] = None,
    desde: date = Query(default_factory=lambda: date.today() - timedelta(days=15)),
    hasta: date = Query(default_factory=lambda: date.today() + timedelta(days=45)),
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    club = require_club(db, current_user)
    data = _coordination_data(
        db, club, temporada_id, coach_user_id,
        desde, hasta,
    )
    season = db.get(Temporada, temporada_id)
    season_name = season.nombre if season else str(temporada_id)
    safe_season = "".join(
        character if character.isalnum() or character in "-_" else "-"
        for character in season_name
    ).strip("-") or str(temporada_id)
    filename = f"calendario_club_{safe_season}.xlsx"
    return StreamingResponse(
        _coordination_workbook(data, season_name, desde, hasta),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        },
    )


@router.put("/entrenadores/{coach_user_id}/visibilidad", response_model=dict)
def set_trainer_visibility(
    coach_user_id: int,
    data: ClubTrainerVisibilityIn,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    club = require_club(db, current_user)
    assignments = db.query(CoachAssignment).filter(
        CoachAssignment.coach_user_id == coach_user_id,
        CoachAssignment.club_id == club.id,
        CoachAssignment.active.is_(True),
    ).all()
    if not assignments:
        raise HTTPException(
            status_code=403,
            detail="El entrenador no pertenece a este club",
        )
    try:
        for assignment in assignments:
            assignment.visible_in_club = data.visible
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail="No se pudo actualizar la visibilidad del entrenador",
        ) from exc
    return {"coach_user_id": coach_user_id, "visible": data.visible}